"""数据集服务：上传、解析、导入、质量校验、样本管理、切分、草稿（设计文档第 5 节）。"""
from __future__ import annotations

import hashlib
import io
import json
import os
import zipfile
from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from sqlalchemy.orm import Session

from app import ids
from app.config import get_settings
from app.errors import ApiError, NotFoundError
from app.models import DatasetQualityReport, DatasetSplit, DatasetVersion, Project, Upload
from app.router_core.normalization import NORMALIZATION_VERSION, normalized_hash
from app.router_core.splitting import DEFAULT_RATIOS, group_split
from app.router_core.taxonomy import LABEL_SCHEMA_VERSION, LABELS
from app.services import artifact_service

ALLOWED_EXTENSIONS = {"csv", "jsonl", "xlsx", "txt"}
ALLOWED_EXTENSIONS_LABEL = {"csv", "jsonl", "xlsx", "txt"}
ALLOWED_MIME_PREFIXES = ("text/", "application/")
PREVIEW_ROWS = 100
SUGGESTED_TEXT_COLUMNS = ("text", "query", "问题", "文本", "query文本", "内容")
SUGGESTED_LABEL_COLUMNS = ("label", "intent", "标签", "意图", "category")

PARQUET_COLUMNS = [
    "sample_id",
    "text",
    "label",
    "group_id",
    "context",
    "source",
    "is_hard_negative",
    "risk_slice",
    "metadata_json",
    "normalized_hash",
]


# ---------------------------------------------------------------- upload
class StreamingUploadWriter:
    """V2 §4.4：上传流式落盘。

    分块写入上传目录内的随机临时文件并增量计算 SHA-256；超过大小限制
    立即终止并删除临时文件；完成后原子移动（os.replace）到最终路径。
    服务端内存占用与文件大小无关。
    """

    def __init__(self, db: Session, project_id: str, original_name: str, content_type: str | None) -> None:
        project = db.get(Project, project_id)
        if project is None:
            raise NotFoundError("Project", project_id)
        ext = Path(original_name).suffix.lstrip(".").lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise ApiError(
                "UNSUPPORTED_FILE_TYPE",
                f"不支持的文件类型 .{ext}，允许: {sorted(ALLOWED_EXTENSIONS)}",
                400,
            )
        self._project_id = project_id
        self._original_name = original_name
        self._content_type = content_type
        self._ext = ext
        self._settings = get_settings()
        self._limit = self._settings.max_upload_mb * 1024 * 1024
        self._settings.uploads_dir.mkdir(parents=True, exist_ok=True)
        self._tmp_path = self._settings.uploads_dir / f".tmp-{ids.prefixed(ids.UPLOAD)}.{ext}"
        self._fh = open(self._tmp_path, "wb")
        self._hasher = hashlib.sha256()
        self._size = 0
        self._aborted = False

    def write(self, chunk: bytes) -> None:
        if self._aborted:
            raise ApiError("UPLOAD_ABORTED", "上传已终止", 400)
        self._size += len(chunk)
        if self._size > self._limit:
            self.abort()
            raise ApiError("FILE_TOO_LARGE", f"文件超过限制 {self._settings.max_upload_mb}MB", 400)
        self._fh.write(chunk)
        self._hasher.update(chunk)

    def finish(self, db: Session) -> Upload:
        upload_id = ids.prefixed(ids.UPLOAD)
        final_path = self._settings.uploads_dir / f"{upload_id}.{self._ext}"
        self._fh.close()
        os.replace(self._tmp_path, final_path)
        upload = Upload(
            id=upload_id,
            project_id=self._project_id,
            original_name=self._original_name[:500],
            safe_path=str(final_path),
            sha256=self._hasher.hexdigest(),
            size_bytes=self._size,
            content_type=self._content_type,
            status="PENDING",
        )
        db.add(upload)
        db.commit()
        return upload

    def abort(self) -> None:
        self._aborted = True
        try:
            if not self._fh.closed:
                self._fh.close()
        finally:
            self._tmp_path.unlink(missing_ok=True)


def save_upload(db: Session, project_id: str, original_name: str, content: bytes, content_type: str | None) -> Upload:
    """字节版上传（CLI / 测试内部调用）；Web 上传走流式 StreamingUploadWriter。"""
    writer = StreamingUploadWriter(db, project_id, original_name, content_type)
    try:
        writer.write(content)
        return writer.finish(db)
    except Exception:
        writer.abort()
        raise


def _guard_xlsx(content: bytes, settings) -> None:
    """V2 §4.4：XLSX 压缩炸弹防护——解压后总大小 / sheet 数 / 首表行列上限。"""
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            expanded = sum(info.file_size for info in archive.infolist())
    except zipfile.BadZipFile as exc:
        raise ApiError("PARSE_ERROR", f"XLSX 解析失败: {exc}", 400) from exc
    if expanded > settings.max_xlsx_expand_mb * 1024 * 1024:
        raise ApiError(
            "ARCHIVE_EXPANSION_TOO_LARGE",
            f"XLSX 解压后 {expanded // (1024 * 1024)}MB，超过上限 {settings.max_xlsx_expand_mb}MB",
            400,
        )
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True)
    except Exception as exc:
        raise ApiError("PARSE_ERROR", f"XLSX 解析失败: {exc}", 400) from exc
    try:
        if len(workbook.sheetnames) > settings.max_xlsx_sheets:
            raise ApiError("PARSE_ERROR", f"XLSX sheet 数超过上限 {settings.max_xlsx_sheets}", 400)
        first = workbook[workbook.sheetnames[0]]
        if (first.max_row or 0) > settings.max_xlsx_rows:
            raise ApiError("PARSE_ERROR", f"XLSX 行数超过上限 {settings.max_xlsx_rows}", 400)
        if (first.max_column or 0) > settings.max_xlsx_cols:
            raise ApiError("PARSE_ERROR", f"XLSX 列数超过上限 {settings.max_xlsx_cols}", 400)
    finally:
        workbook.close()


def _decode_bytes(content: bytes, encoding: str | None) -> tuple[str, str]:
    """返回 (文本, 实际编码)。依次尝试 utf-8-sig / utf-8 / gbk / gb18030。"""
    if encoding:
        try:
            return content.decode(encoding), encoding
        except (UnicodeDecodeError, LookupError) as exc:
            raise ApiError("ENCODING_ERROR", f"使用指定编码 {encoding} 解码失败", 400) from exc
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return content.decode(enc), enc
        except UnicodeDecodeError:
            continue
    raise ApiError("ENCODING_ERROR", "文件无法以 UTF-8/GBK 解码，请指定编码", 400)


def read_tabular(upload: Upload, encoding: str | None = None) -> tuple[pd.DataFrame, str]:
    """解析上传文件为 DataFrame（全部列保持字符串）。"""
    settings = get_settings()
    content = Path(upload.safe_path).read_bytes()
    ext = Path(upload.safe_path).suffix.lstrip(".").lower()

    if ext == "xlsx":
        _guard_xlsx(content, settings)
        try:
            df = pd.read_excel(io.BytesIO(content), sheet_name=0, dtype=str)
        except Exception as exc:
            raise ApiError("PARSE_ERROR", f"XLSX 解析失败: {exc}", 400) from exc
        used_encoding = "binary"
    elif ext == "csv":
        text, used_encoding = _decode_bytes(content, encoding)
        try:
            df = pd.read_csv(io.StringIO(text), dtype=str, keep_default_na=False, sep=None, engine="python")
        except Exception as exc:
            raise ApiError("PARSE_ERROR", f"CSV 解析失败: {exc}", 400) from exc
    elif ext == "jsonl":
        text, used_encoding = _decode_bytes(content, encoding)
        rows = []
        for line_no, line in enumerate(text.splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ApiError("PARSE_ERROR", f"JSONL 第 {line_no} 行解析失败: {exc}", 400) from exc
            if not isinstance(obj, dict):
                raise ApiError("PARSE_ERROR", f"JSONL 第 {line_no} 行不是对象", 400)
            rows.append({k: str(v) if v is not None else None for k, v in obj.items()})
        df = pd.DataFrame(rows)
    elif ext == "txt":
        text, used_encoding = _decode_bytes(content, encoding)
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        df = pd.DataFrame({"text": lines})
    else:
        raise ApiError("UNSUPPORTED_FILE_TYPE", f"不支持的扩展名 {ext}", 400)

    if len(df) > settings.max_rows_per_file:
        raise ApiError("ROWS_EXCEEDED", f"行数 {len(df)} 超过上限 {settings.max_rows_per_file}", 400)

    df = df.fillna("")
    df = df.map(lambda v: v.strip() if isinstance(v, str) else v)
    return df, used_encoding


def preview_upload(db: Session, upload_id: str, encoding: str | None = None) -> dict:
    upload = db.get(Upload, upload_id)
    if upload is None:
        raise NotFoundError("Upload", upload_id)
    df, used_encoding = read_tabular(upload, encoding)
    columns = [str(c) for c in df.columns.tolist()]

    def _suggest(candidates: tuple[str, ...]) -> str | None:
        for cand in candidates:
            for col in columns:
                if col.lower() == cand:
                    return col
        for cand in candidates:
            for col in columns:
                if cand in col.lower():
                    return col
        return None

    rows = df.head(PREVIEW_ROWS).to_dict(orient="records")
    return {
        "upload_id": upload.id,
        "original_name": upload.original_name,
        "columns": columns,
        "row_count": int(len(df)),
        "used_encoding": used_encoding,
        "rows": [{k: (v if v != "" else None) for k, v in row.items()} for row in rows],
        "suggested_columns": {
            "text": _suggest(SUGGESTED_TEXT_COLUMNS),
            "label": _suggest(SUGGESTED_LABEL_COLUMNS),
        },
    }


# ---------------------------------------------------------------- import
def import_upload(db: Session, upload_id: str, config: dict) -> DatasetVersion:
    """按列映射与导入配置执行导入（设计文档 5.3）。"""
    upload = db.get(Upload, upload_id)
    if upload is None:
        raise NotFoundError("Upload", upload_id)

    mode = config.get("mode", "prelabeled")
    if mode not in ("prelabeled", "unlabeled", "single_label"):
        raise ApiError("VALIDATION_ERROR", f"未知导入模式 {mode}", 422)
    columns = config.get("columns") or {}
    text_col = columns.get("text")
    if not text_col:
        raise ApiError("VALIDATION_ERROR", "缺少文本列映射", 422)
    label_col = columns.get("label")
    if mode == "prelabeled" and not label_col:
        raise ApiError("VALIDATION_ERROR", "已标注数据导入必须映射标签列", 422)
    default_label = config.get("default_label")
    if mode == "single_label" and default_label not in LABELS:
        raise ApiError("VALIDATION_ERROR", f"按标签导入需要合法标签，得到 {default_label!r}", 422)

    df, _encoding = read_tabular(upload, config.get("encoding"))
    if text_col not in df.columns:
        raise ApiError("VALIDATION_ERROR", f"文本列 {text_col} 不存在", 422)
    if label_col and label_col not in df.columns:
        raise ApiError("VALIDATION_ERROR", f"标签列 {label_col} 不存在", 422)

    label_mapping = config.get("label_mapping") or {}

    records: list[dict] = []
    errors: list[dict] = []
    warnings: list[dict] = []

    seen_hash_label: dict[str, str] = {}
    duplicate_count = 0
    empty_rows: list[int] = []
    too_long_rows: list[int] = []
    bad_labels: dict[str, int] = {}
    settings = get_settings()

    for row_no, row in df.iterrows():
        raw_text = row.get(text_col)
        text = str(raw_text).strip() if raw_text is not None else ""
        if not text:
            empty_rows.append(int(row_no) + 1)
            continue
        if len(text) > settings.max_text_chars:
            too_long_rows.append(int(row_no) + 1)
            continue

        # 标签解析
        label: str | None = None
        if mode == "single_label":
            label = default_label
        elif mode == "prelabeled" and label_col:
            raw_label = row.get(label_col)
            raw_label = str(raw_label).strip() if raw_label is not None else ""
            mapped = label_mapping.get(raw_label, raw_label)
            if mapped in ("", "__skip__", None):
                continue
            if mapped not in LABELS:
                bad_labels[mapped] = bad_labels.get(mapped, 0) + 1
                continue
            label = mapped

        context = _cell(row, columns.get("context"))
        group_id = _cell(row, columns.get("group_id"))
        source = _cell(row, columns.get("source")) or upload.original_name
        risk_slice = _cell(row, columns.get("risk_slice"))
        hard_raw = _cell(row, columns.get("is_hard_negative"))
        is_hard = hard_raw.lower() in ("true", "1", "yes", "是", "y") if hard_raw else False

        n_hash = normalized_hash(text, context)
        if label is not None:
            if n_hash in seen_hash_label and seen_hash_label[n_hash] != label:
                errors.append(
                    {
                        "code": "DATASET_LABEL_CONFLICT",
                        "message": "同一规范化文本对应多个标签",
                        "details": {"text_hash": n_hash[:16], "labels": [seen_hash_label[n_hash], label]},
                    }
                )
            else:
                seen_hash_label[n_hash] = label
        records.append(
            {
                "sample_id": ids.prefixed(ids.SAMPLE),
                "text": text,
                "label": label,
                "group_id": group_id,
                "context": context,
                "source": source,
                "is_hard_negative": is_hard,
                "risk_slice": risk_slice,
                "metadata_json": None,
                "normalized_hash": n_hash,
                "_row_no": int(row_no) + 1,
            }
        )

    if empty_rows:
        errors.append(
            {
                "code": "EMPTY_TEXT",
                "message": f"{len(empty_rows)} 行文本为空或仅包含空白，已跳过",
                "details": {"rows": empty_rows[:20]},
            }
        )
    if too_long_rows:
        errors.append(
            {
                "code": "TEXT_TOO_LONG",
                "message": f"{len(too_long_rows)} 行超过 {settings.max_text_chars} 字符，已跳过",
                "details": {"rows": too_long_rows[:20]},
            }
        )
    if bad_labels:
        errors.append(
            {
                "code": "INVALID_LABEL",
                "message": "存在不在 Label Schema 内的标签",
                "details": {"labels": bad_labels},
            }
        )

    # 完全重复（同 hash 同标签）去重，保留首条
    unique_records: list[dict] = []
    seen_exact: set[str] = set()
    for rec in records:
        key = rec["normalized_hash"] + "|" + str(rec["label"])
        if rec["label"] is not None and key in seen_exact:
            duplicate_count += 1
            continue
        if rec["label"] is not None:
            seen_exact.add(key)
        unique_records.append(rec)

    if not unique_records:
        raise ApiError("PARSE_ERROR", "导入后没有有效样本", 400)

    dup_rate = duplicate_count / max(len(records), 1)
    if dup_rate > 0.10:
        warnings.append(
            {"code": "HIGH_DUPLICATE_RATE", "message": f"重复率 {dup_rate:.1%} 超过 10%", "details": {"duplicates": duplicate_count}}
        )

    label_counts: dict[str, int] = {}
    for rec in unique_records:
        if rec["label"]:
            label_counts[rec["label"]] = label_counts.get(rec["label"], 0) + 1
    labeled = sum(label_counts.values())
    unlabeled = len(unique_records) - labeled
    if mode == "prelabeled" and unlabeled > 0:
        warnings.append({"code": "UNLABELED_ROWS", "message": f"{unlabeled} 条样本无标签", "details": {}})
    for lab, count in label_counts.items():
        if count < 20:
            warnings.append({"code": "CLASS_TOO_FEW", "message": f"类别 {lab} 样本数 {count} 少于 20", "details": {"label": lab, "count": count}})
    if label_counts:
        ratio = max(label_counts.values()) / max(min(label_counts.values()), 1)
        if ratio > 10:
            warnings.append({"code": "CLASS_IMBALANCE", "message": f"类别最大/最小样本数比 {ratio:.1f} 超过 10", "details": {}})
    has_group = sum(1 for r in unique_records if r["group_id"])
    if has_group / len(unique_records) < 0.5:
        warnings.append({"code": "GROUP_ID_MISSING", "message": f"{len(unique_records) - has_group} 条样本缺少 group_id", "details": {}})
    hard_count = sum(1 for r in unique_records if r["is_hard_negative"])
    if labeled and hard_count / labeled < 0.20:
        warnings.append({"code": "HARD_NEGATIVE_LOW", "message": f"hard negative 占比 {hard_count / labeled:.1%} 低于 20%", "details": {"count": hard_count}})

    frame = pd.DataFrame(unique_records)
    frame = frame[PARQUET_COLUMNS]

    # 写入制品目录
    dataset_id = ids.prefixed(ids.DATASET)
    ds_dir = artifact_service.dataset_dir(upload.project_id, dataset_id)
    parquet_path = ds_dir / "data.parquet"
    _write_parquet(frame, parquet_path)

    status = "FROZEN" if mode == "prelabeled" else "DRAFT"
    parent_version = (
        db.query(DatasetVersion)
        .filter(DatasetVersion.project_id == upload.project_id)
        .count()
    ) + 1

    dataset = DatasetVersion(
        id=dataset_id,
        project_id=upload.project_id,
        parent_id=None,
        version=parent_version,
        name=config.get("name") or upload.original_name,
        origin="import",
        status=status,
        parquet_path=str(parquet_path),
        raw_path=upload.safe_path,
        sample_count=len(frame),
        labeled_count=labeled,
        label_distribution=label_counts,
        manifest={
            "normalization_version": NORMALIZATION_VERSION,
            "label_schema_version": LABEL_SCHEMA_VERSION,
            "source_upload_id": upload.id,
            "raw_sha256": upload.sha256,
            "import_config": {"mode": mode, "columns": columns, "label_mapping": label_mapping, "default_label": default_label},
            "created_at": datetime.now(UTC).isoformat(),
        },
        change_summary=config.get("change_summary", f"导入自 {upload.original_name}"),
    )
    db.add(dataset)
    db.flush()  # 先落 DatasetVersion 行，保证质量报告外键可满足

    report = build_report(errors, warnings, frame)
    db.add(DatasetQualityReport(id=ids.prefixed("qar"), dataset_id=dataset_id, report_json=report))

    upload.status = "IMPORTED"
    db.commit()
    return dataset


def _cell(row: pd.Series, col: str | None) -> str | None:
    if not col:
        return None
    val = row.get(col)
    if val is None:
        return None
    sval = str(val).strip()
    return sval or None


def build_report(errors: list[dict], warnings: list[dict], frame: pd.DataFrame) -> dict:
    label_counts = frame["label"].value_counts().to_dict() if "label" in frame.columns else {}
    return {
        "errors": errors,
        "warnings": warnings,
        "stats": {
            "rows": int(len(frame)),
            "labeled": int(frame["label"].notna().sum()) if "label" in frame.columns else 0,
            "unlabeled": int(frame["label"].isna().sum()) if "label" in frame.columns else 0,
            "label_distribution": {str(k): int(v) for k, v in label_counts.items()},
            "unique_hashes": int(frame["normalized_hash"].nunique()) if "normalized_hash" in frame.columns else 0,
            "has_group_id": int((frame["group_id"].notna() & (frame["group_id"] != "")).sum()) if "group_id" in frame.columns else 0,
            "hard_negative": int(frame["is_hard_negative"].sum()) if "is_hard_negative" in frame.columns else 0,
            "max_label_support": int(max(label_counts.values())) if label_counts else 0,
            "min_label_support": int(min(label_counts.values())) if label_counts else 0,
            "non_write_support": int(sum(v for k, v in label_counts.items() if k != "write_action")),
        },
        "generated_at": datetime.now(UTC).isoformat(),
    }


# ---------------------------------------------------------------- samples
def load_dataset_frame(dataset: DatasetVersion) -> pd.DataFrame:
    path = dataset.parquet_path
    if not Path(path).is_file():
        raise ApiError("ARTIFACT_INCOMPLETE", "数据集 parquet 文件缺失", 409, {"path": Path(path).name})
    return pd.read_parquet(path)


def _write_parquet(frame: pd.DataFrame, path: Path) -> None:
    table = pa.Table.from_pandas(frame, preserve_index=False)
    pq.write_table(table, path)


def list_samples(db: Session, dataset_id: str, filters: dict, page: int = 1, page_size: int = 50) -> dict:
    dataset = db.get(DatasetVersion, dataset_id)
    if dataset is None:
        raise NotFoundError("Dataset", dataset_id)
    frame = load_dataset_frame(dataset)

    query_text = (filters.get("q") or "").strip()
    label_filter = filters.get("label")
    only_unlabeled = bool(filters.get("unlabeled_only"))

    mask = pd.Series(True, index=frame.index)
    if query_text:
        mask &= frame["text"].str.contains(query_text, case=False, regex=False)
    if only_unlabeled:
        mask &= frame["label"].isna() | (frame["label"] == "")
    elif label_filter:
        mask &= frame["label"] == label_filter

    total = int(mask.sum())
    start = (page - 1) * page_size
    rows = frame[mask].iloc[start : start + page_size]

    return {
        "dataset_id": dataset_id,
        "total": total,
        "page": page,
        "page_size": page_size,
        "samples": _samples_to_dicts(rows),
    }


def _samples_to_dicts(rows: pd.DataFrame) -> list[dict]:
    out = []
    for _, row in rows.iterrows():
        out.append(
            {
                "sample_id": row["sample_id"],
                "text": row["text"],
                "label": row["label"] if row["label"] not in ("", None) and not pd.isna(row["label"]) else None,
                "group_id": row.get("group_id") if not pd.isna(row.get("group_id")) else None,
                "context": row.get("context") if not pd.isna(row.get("context")) else None,
                "source": row.get("source") if not pd.isna(row.get("source")) else None,
                "is_hard_negative": bool(row.get("is_hard_negative")) if not pd.isna(row.get("is_hard_negative")) else False,
                "risk_slice": row.get("risk_slice") if not pd.isna(row.get("risk_slice")) else None,
                "normalized_hash": row["normalized_hash"],
            }
        )
    return out


def update_sample(db: Session, dataset_id: str, sample_id: str, patch: dict) -> dict:
    """仅 DRAFT 数据集可改样本；改标签时做冲突检查。"""
    dataset = db.get(DatasetVersion, dataset_id)
    if dataset is None:
        raise NotFoundError("Dataset", dataset_id)
    if dataset.status != "DRAFT":
        raise ApiError("DATASET_IMMUTABLE", "已冻结数据集不可修改，请创建下一版本草稿", 409)

    frame = load_dataset_frame(dataset)
    idx = frame.index[frame["sample_id"] == sample_id]
    if len(idx) == 0:
        raise NotFoundError("Sample", sample_id)
    pos = idx[0]

    new_label = patch.get("label")
    if new_label is not None and new_label not in LABELS:
        raise ApiError("INVALID_LABEL", f"非法标签 {new_label}", 422)

    if new_label is not None:
        n_hash = frame.at[pos, "normalized_hash"]
        conflict = frame[(frame["normalized_hash"] == n_hash) & (frame.index != pos) & (frame["label"].notna()) & (frame["label"] != new_label)]
        if len(conflict) > 0:
            raise ApiError(
                "DATASET_LABEL_CONFLICT",
                "同一规范化文本在其他样本上已有不同标签",
                409,
                {"sample_ids": conflict["sample_id"].head(10).tolist()},
            )
        frame.at[pos, "label"] = new_label

    if "is_hard_negative" in patch:
        frame.at[pos, "is_hard_negative"] = bool(patch["is_hard_negative"])
    if "risk_slice" in patch:
        frame.at[pos, "risk_slice"] = patch["risk_slice"] or None
    if "group_id" in patch:
        frame.at[pos, "group_id"] = patch["group_id"] or None
    if "note" in patch:
        meta = json.loads(frame.at[pos, "metadata_json"]) if frame.at[pos, "metadata_json"] else {}
        meta["note"] = patch["note"]
        frame.at[pos, "metadata_json"] = json.dumps(meta, ensure_ascii=False)

    _write_parquet(frame, Path(dataset.parquet_path))
    _refresh_stats(db, dataset, frame)
    db.commit()
    return _samples_to_dicts(frame.iloc[[pos]])[0]


def _refresh_stats(db: Session, dataset: DatasetVersion, frame: pd.DataFrame | None = None) -> None:
    frame = frame if frame is not None else load_dataset_frame(dataset)
    labeled_series = frame["label"].fillna("")
    dataset.sample_count = int(len(frame))
    dataset.labeled_count = int((labeled_series != "").sum())
    dataset.label_distribution = {str(k): int(v) for k, v in labeled_series.value_counts().items() if k}


def validate_dataset(db: Session, dataset_id: str) -> dict:
    """重新执行质量检查并存储报告。"""
    dataset = db.get(DatasetVersion, dataset_id)
    if dataset is None:
        raise NotFoundError("Dataset", dataset_id)
    frame = load_dataset_frame(dataset)
    errors: list[dict] = []
    warnings: list[dict] = []

    # V2 §3.4：非法标签 / 空标签阻断，并指出样本 ID
    bad_mask = frame["label"].notna() & (frame["label"] != "") & ~frame["label"].isin(LABELS)
    if bad_mask.any():
        errors.append(
            {
                "code": "INVALID_LABEL",
                "message": f"{int(bad_mask.sum())} 条样本标签不在五分类内",
                "details": {
                    "sample_ids": frame.loc[bad_mask, "sample_id"].head(10).tolist(),
                    "labels": sorted(set(frame.loc[bad_mask, "label"])),
                },
            }
        )
    empty_mask = frame["label"] == ""
    if empty_mask.any():
        errors.append(
            {
                "code": "EMPTY_LABEL",
                "message": f"{int(empty_mask.sum())} 条样本标签为空字符串（应删除或补标）",
                "details": {"sample_ids": frame.loc[empty_mask, "sample_id"].head(10).tolist()},
            }
        )

    hash_labels: dict[str, set] = {}
    for _, row in frame.iterrows():
        if row["label"]:
            hash_labels.setdefault(row["normalized_hash"], set()).add(row["label"])
    for n_hash, labs in hash_labels.items():
        if len(labs) > 1:
            conflict_ids = frame.loc[frame["normalized_hash"] == n_hash, "sample_id"].head(10).tolist()
            errors.append(
                {
                    "code": "DATASET_LABEL_CONFLICT",
                    "message": "同一规范化文本对应多个标签",
                    "details": {"text_hash": n_hash[:16], "labels": sorted(labs), "sample_ids": conflict_ids},
                }
            )
    present = {lab for lab in hash_labels.values() for lab in lab}
    if present:
        missing_classes = [lab for lab in LABELS if lab not in present]
        if missing_classes:
            errors.append(
                {
                    "code": "MISSING_LABEL_CLASS",
                    "message": f"缺少类别 {missing_classes}：五分类契约要求冻结前五类齐全",
                    "details": {"missing": missing_classes, "present": sorted(present)},
                }
            )
    if len(errors) == 0 and (frame["label"].isna() | (frame["label"] == "")).all():
        errors.append({"code": "NO_LABELS", "message": "数据集没有任何标签", "details": {}})

    report = build_report(errors, warnings, frame)
    db.add(DatasetQualityReport(id=ids.prefixed("qar"), dataset_id=dataset_id, report_json=report))
    db.commit()
    return report


def latest_report(db: Session, dataset_id: str) -> dict | None:
    row = (
        db.query(DatasetQualityReport)
        .filter(DatasetQualityReport.dataset_id == dataset_id)
        .order_by(DatasetQualityReport.created_at.desc())
        .first()
    )
    return row.report_json if row else None


# ---------------------------------------------------------------- split
def create_split(db: Session, dataset_id: str, ratios: dict | None = None, seed: int = 42) -> DatasetSplit:
    dataset = db.get(DatasetVersion, dataset_id)
    if dataset is None:
        raise NotFoundError("Dataset", dataset_id)
    frame = load_dataset_frame(dataset)
    unlabeled = frame["label"].isna() | (frame["label"] == "")
    if unlabeled.any():
        raise ApiError("UNLABELED_SAMPLES", f"仍有 {int(unlabeled.sum())} 条未标注样本，无法切分", 409)

    result = group_split(frame, ratios=ratios or DEFAULT_RATIOS, seed=seed)
    split_frame = pd.DataFrame(
        {
            "sample_id": frame["sample_id"],
            "split": result.df["split"],
            "is_risk_test": result.df["is_risk_test"],
        }
    )

    split_id = ids.prefixed(ids.SPLIT)
    ds_dir = Path(dataset.parquet_path).parent
    split_path = ds_dir / f"split_{split_id}.parquet"
    _write_parquet(split_frame, split_path)

    stats = dict(result.stats)
    stats["warnings"] = result.warnings
    split = DatasetSplit(
        id=split_id,
        dataset_id=dataset_id,
        seed=seed,
        ratios=ratios or DEFAULT_RATIOS,
        parquet_path=str(split_path),
        stats_json=stats,
    )
    db.add(split)
    db.commit()
    return split


def load_split_frame(split: DatasetSplit) -> pd.DataFrame:
    return pd.read_parquet(split.parquet_path)


# ---------------------------------------------------------------- draft & commit
def create_draft(db: Session, source_dataset_id: str, changes: list[dict], name: str | None = None) -> DatasetVersion:
    """从任意版本创建下一版本草稿（用于错误样本回流 / 主动学习）。"""
    source = db.get(DatasetVersion, source_dataset_id)
    if source is None:
        raise NotFoundError("Dataset", source_dataset_id)

    frame = load_dataset_frame(source)
    summary: list[str] = []

    for change in changes:
        action = change.get("action")
        if action == "update":
            idx = frame.index[frame["sample_id"] == change.get("sample_id")]
            if len(idx) == 0:
                raise ApiError("VALIDATION_ERROR", f"样本不存在: {change.get('sample_id')}", 422)
            pos = idx[0]
            if "label" in change and change["label"]:
                if change["label"] not in LABELS:  # V2 §3.4：update 与 add 同样校验
                    raise ApiError("INVALID_LABEL", f"非法标签 {change['label']}", 422)
                frame.at[pos, "label"] = change["label"]
            if "is_hard_negative" in change:
                frame.at[pos, "is_hard_negative"] = bool(change["is_hard_negative"])
            if "risk_slice" in change:
                frame.at[pos, "risk_slice"] = change.get("risk_slice")
            summary.append(f"更新 {change.get('sample_id')}")
        elif action == "add":
            text = str(change.get("text", "")).strip()
            if not text:
                raise ApiError("VALIDATION_ERROR", "新增样本缺少文本", 422)
            label = change.get("label")
            if label and label not in LABELS:
                raise ApiError("INVALID_LABEL", f"非法标签 {label}", 422)
            context = change.get("context")
            frame = pd.concat(
                [
                    frame,
                    pd.DataFrame(
                        [
                            {
                                "sample_id": ids.prefixed(ids.SAMPLE),
                                "text": text,
                                "label": label,
                                "group_id": change.get("group_id"),
                                "context": context,
                                "source": change.get("source") or "error_feedback",
                                "is_hard_negative": bool(change.get("is_hard_negative", False)),
                                "risk_slice": change.get("risk_slice"),
                                "metadata_json": json.dumps({"note": change.get("note")}, ensure_ascii=False) if change.get("note") else None,
                                "normalized_hash": normalized_hash(text, context),
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )
            summary.append(f"新增样本: {text[:30]}")
        elif action == "remove":
            frame = frame[frame["sample_id"] != change.get("sample_id")]
            summary.append(f"删除 {change.get('sample_id')}")
        else:
            raise ApiError("VALIDATION_ERROR", f"未知变更类型 {action}", 422)

    dataset_id = ids.prefixed(ids.DATASET)
    ds_dir = artifact_service.dataset_dir(source.project_id, dataset_id)
    parquet_path = ds_dir / "data.parquet"
    _write_parquet(frame[PARQUET_COLUMNS], parquet_path)

    draft = DatasetVersion(
        id=dataset_id,
        project_id=source.project_id,
        parent_id=source.id,
        version=_next_version(db, source.project_id),
        name=name or f"{source.name} 草稿",
        origin="draft",
        status="DRAFT",
        parquet_path=str(parquet_path),
        sample_count=len(frame),
        manifest={**(source.manifest or {}), "parent_id": source.id, "created_at": datetime.now(UTC).isoformat()},
        change_summary="; ".join(summary[:50]) or "无变更",
    )
    _refresh_stats(db, draft, frame)
    db.add(draft)
    db.commit()
    return draft


def _next_version(db: Session, project_id: str) -> int:
    latest = (
        db.query(DatasetVersion)
        .filter(DatasetVersion.project_id == project_id)
        .order_by(DatasetVersion.version.desc())
        .first()
    )
    return (latest.version + 1) if latest else 1


def commit_draft(db: Session, draft_id: str) -> DatasetVersion:
    draft = db.get(DatasetVersion, draft_id)
    if draft is None:
        raise NotFoundError("DatasetDraft", draft_id)
    if draft.status != "DRAFT":
        raise ApiError("DATASET_IMMUTABLE", "只有 DRAFT 状态可以提交冻结", 409)

    report = validate_dataset(db, draft_id)
    if report["errors"]:
        db.commit()
        raise ApiError("QUALITY_ERRORS", "存在阻断级数据问题，无法冻结", 409, {"report": report["errors"]})

    draft.status = "FROZEN"
    db.commit()
    db.refresh(draft)
    return draft
